import openpyxl
from company import Company
from meachine import Meachine

class ReadFileExcel:

    def __init__(self):
        pass
    
    # list cac may
    def docMayVaCongTy(self, name):
        wb = openpyxl.load_workbook(name);
        sheet = wb.active;
        ans = [];
        getMaCty = [];
        getMaMay = [];
        for row in sheet.values:
            a = str(row[1]);
            b = str(row[5]);
            if(a not in getMaCty):
                getMaCty.append(a);
                ans.append(Company(a, str(row[2])));
            if(b not in getMaMay):
                getMaMay.append(b);
                for i in range(0, len(ans)):
                    if(ans[i].maCty in a):
                        ans[i].upDateMeachine(Meachine(b, row[6], row[8], str(row[10])))
        
        wb.close();
        return ans;
        

    # them tien hang thang vao cac may
    def tinhTienTheoThang(self, T, sheet):
        CtyVaMay = [];
        for i in T:
            tmp = Company(i.maCty, i.tenCongTy);
            for j in i.meachines:
                tmp.meachines.append(Meachine(j.maMay, j.loaiMay, j.diaChi, j.soDienThoai))
            CtyVaMay.append(tmp);
        for row in sheet.values:
            count = 0;
            check = 0;
            mamay = "";
            for cell in row:
                if(str(cell) != "None"):
                    count = count + 1;
                if(count == 1):
                    for i in range(0, len(CtyVaMay)):
                        for j in range(0, len(CtyVaMay[i].meachines)):
                            if(CtyVaMay[i].meachines[j].maMay == str(cell)):
                                mamay = str(cell);
                                check = 1;
                if(check == 0 and count >= 1):
                    break;
                if(count == 13 and str(cell) != "None"):
                    for i in range(0, len(CtyVaMay)):
                        for j in range(0, len(CtyVaMay[i].meachines)):
                            if(CtyVaMay[i].meachines[j].maMay == mamay):
                                CtyVaMay[i].meachines[j].upDateMoney(float(cell))
                
                if(count == 14 and str(cell) != "None"):
                    for i in range(0, len(CtyVaMay)):
                        for j in range(0, len(CtyVaMay[i].meachines)):
                            if(CtyVaMay[i].meachines[j].maMay == mamay):
                                CtyVaMay[i].meachines[j].upDatePhi(float(cell))

                for i in range(0, len(CtyVaMay)):    
                    CtyVaMay[i].upDateTienVaPhiCty()
        return CtyVaMay;

    def tinhTien(self, tmp, name):
        wb = openpyxl.load_workbook(name);
        ans = [];
        name_ans = [];
        for i in wb.sheetnames:
            a = wb[i]
            name_ans.append(i);
            # print(a)
            ans.append(self.tinhTienTheoThang(tmp, a));
        wb.close();
        return ans, name_ans;


# a = ReadFileExcel();
# b = a.docMayVaCongTy("../myexcel/source.xlsx");
# c = a.tinhTien(b, "../myexcel/follow.xlsx")
# for i in c:
#     for j in i:
#         print(j.maCty + " " + str(j.tienCongTy))
