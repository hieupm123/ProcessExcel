import openpyxl
from company import Company
from meachine import Meachine
from readFileExcel import ReadFileExcel

class WriteFileExcel:

    def __init__(self):
        self.wb = openpyxl.Workbook();
        self.wb.remove_sheet(self.wb["Sheet"])
    
    def getPhanTram(self, a, b):
        if(b == 0 and a > 0):
            return "100%";
        if(b == 0 and a < 0):
            return "-100%";
        if(b == 0 and a == 0):
            return "0%"
        c = (float(a) / float(b)) - 1;
        c = c * 100;
        c = round(c, 2)
        return str(c) + "%";
    

    def luuFile(self, name):
        self.wb.save(name);
    
    def taoSheetMoi(self, name):
        if(name in self.wb.sheetnames):
            self.wb.remove_sheet(self.wb[name])

        self.wb.create_sheet(name);
        return self.wb[name];

    def danhSachDVCNT(self, inMonth, danhsach):
        sheet = self.taoSheetMoi("danhsachDVCNT")
        a = ("STT", "Mã Công Ty", "Tên Công Ty", "Mã Máy", 
             "Loại Máy", "Địa chỉ", "Số điện thoại", "Tổng doanh thu", "Tổng phí");
        sheet.append(a);
        b = {};
        c = {};
        for i in inMonth:
            for j in i:
                for k in j.meachines:
                    if(k.maMay not in b):
                        b.update({str(k.maMay) : float(k.tienMay)});
                        c.update({str(k.maMay) : float(k.phiDichVu)})
                    else:
                        b[k.maMay] += k.tienMay;
                        c[k.maMay] += k.phiDichVu;
                        

        cnt = 0;
        for i in danhsach:
            for j in i.meachines:
                if(j.maMay in b and b[j.maMay] != 0):
                    cnt = cnt + 1;
                    sheet.append((cnt, i.maCty, i.tenCongTy, j.maMay,
                                  j.loaiMay, j.diaChi, j.soDienThoai, b[j.maMay],
                                  c[j.maMay]));


    def doanhSoChiTiet(self):
        sheet = self.taoSheetMoi("doanhSoChiTiet")

    def baoCaoChiTiet(self, thang, inMonth, thangBao, maCongTy):
        sheet = self.taoSheetMoi("BCCT " + thangBao)
        sheet.append(("STT", "TID", "Loại máy", "Doanh số giao dịch", "", "Doanh thu phí", "", "Địa chỉ", "Số điện thoại"))
        sheet.append(("", "", "", "Doanh số GD trong kỳ", "Tăng trưởng so với tháng trước",
                      "Doanh thu trong kì", "Tăng trưởng so với tháng trước", "", ""))
        sheet.merge_cells("A1:A2")
        sheet.merge_cells("B1:B2")
        sheet.merge_cells("C1:C2")
        sheet.merge_cells("D1:E1")
        sheet.merge_cells("F1:G1")
        sheet.merge_cells("H1:H2")
        sheet.merge_cells("I1:I2")
        a = 0; b = 0;
        for i in range(0, len(thang)):
            if(thangBao in thang[i]):
                a = i; b = i - 1;
        cnt = 0;
        for i in range(0, len(inMonth[a])):
            if(inMonth[a][i].maCty not in maCongTy):
                continue
            for j in range(0,len(inMonth[a][i].meachines)):
                check = 0;
                for k in range(0, 9):
                    if(str(k) in inMonth[a][i].meachines[j].maMay):
                        check = 1;
                if(check == 0):
                    continue;
                cnt = cnt + 1;
                tmp1 = 0; tmp2 = 0;
                
                if(b != -1):
                    tmp1 = float(inMonth[b][i].meachines[j].tienMay);
                    tmp2 = float(inMonth[b][i].meachines[j].phiDichVu);
                tmp1 = self.getPhanTram(float(inMonth[a][i].meachines[j].tienMay), tmp1);
                tmp2 = self.getPhanTram(float(inMonth[a][i].meachines[j].phiDichVu), tmp2);
                T = inMonth[a][i].meachines[j];
                sheet.append((cnt, T.maMay, T.loaiMay, 
                              T.tienMay, tmp1,
                              T.phiDichVu, tmp2, T.diaChi, T.soDienThoai))

    def baoCaoTongHop(self, thang, inMonth, thangBao):
        sheet = self.taoSheetMoi("BCTH " + thangBao)
        sheet.append(("STT", "Tên ĐVCNT", "MID", "Doanh số giao dịch", "", "Doanh thu phí", ""))
        sheet.append(("", "", "", "Doanh số GD trong kỳ", "Tăng trưởng so với tháng trước",
                      "Doanh thu trong kì", "Tăng trưởng so với tháng trước"))
        sheet.merge_cells("A1:A2")
        sheet.merge_cells("B1:B2")
        sheet.merge_cells("C1:C2")
        sheet.merge_cells("D1:E1")
        sheet.merge_cells("F1:G1")
        a = 0; b = 0;
        for i in range(0, len(thang)):
            if(thangBao in thang[i]):
                a = i; b = i - 1;
        cnt = 0
        for i in range(0, len(inMonth[a])):
            check = 0;
            for j in range(0, 9):
                if(str(j) in inMonth[a][i].maCty):
                    check = 1;
            if(check == 0):
                continue;
            else:
                cnt = cnt + 1;
                tmp1 = 0;
                tmp2 = 0;
                if(b != -1):
                    tmp1 = float(inMonth[b][i].tienCongTy);
                    tmp2 = float(inMonth[b][i].phiCongTy);
                tmp1 = self.getPhanTram(float(inMonth[a][i].tienCongTy), tmp1);
                tmp2 = self.getPhanTram(float(inMonth[a][i].phiCongTy), tmp2);
                           
                sheet.append((cnt, inMonth[a][i].tenCongTy, inMonth[a][i].maCty, 
                              inMonth[a][i].tienCongTy, tmp1,
                              inMonth[a][i].phiCongTy, tmp2))
                              



# d = WriteFileExcel();
# a = ReadFileExcel();

# b = a.docMayVaCongTy("../myexcel/source.xlsx");
# c, e = a.tinhTien(b, "../myexcel/follow.xlsx");

# d.danhSachDVCNT(c, b);
# may = ["100107389", "100442102"];
# d.baoCaoTongHop(e, c, "T4");
# d.baoCaoChiTiet(e, c, "T4", maMay=may);

# d.luuFile("../myexcel/output_excel.xlsx")
